from datetime import datetime, timedelta

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app import models
from app.db import Base
from app.reports import build_site_report, report_filename


@pytest.fixture()
def db_session(tmp_path):
    engine = create_engine(f"sqlite:///{tmp_path}/reports_test.db")
    Base.metadata.create_all(bind=engine)
    session = sessionmaker(bind=engine)()
    yield session
    session.close()


def _seed(db):
    site = models.Site(name="Kiran Chenna Prod", base_url="https://example.com")
    account = models.Account(site=site, label="demo", username="u", encrypted_password=b"x")
    db.add_all([site, account])
    db.flush()

    base = datetime(2026, 1, 1, 9, 0, 0)
    runs = [
        (base, models.RunStatus.success, 1000),
        (base + timedelta(minutes=15), models.RunStatus.success, 1000),
        (base + timedelta(minutes=30), models.RunStatus.fail, 2000),
        (base + timedelta(minutes=45), models.RunStatus.fail, 2000),
        (base + timedelta(minutes=60), models.RunStatus.success, 1000),
        (base + timedelta(minutes=75), models.RunStatus.running, None),  # must be excluded entirely
    ]
    for started_at, status, duration_ms in runs:
        finished_at = started_at + timedelta(milliseconds=duration_ms) if duration_ms is not None else None
        db.add(
            models.CheckRun(
                site_id=site.id,
                account_id=account.id,
                status=status,
                started_at=started_at,
                finished_at=finished_at,
                duration_ms=duration_ms,
            )
        )
    db.commit()
    return site


def test_report_has_two_sheets_with_expected_headers(db_session):
    site = _seed(db_session)
    buffer = build_site_report(db_session, site)
    wb = load_workbook(buffer)

    assert wb.sheetnames == ["Check Runs", "Status Changes"]
    assert [c.value for c in wb["Check Runs"][1]] == [
        "Started At",
        "Finished At",
        "Duration (s)",
        "Account",
        "Status",
        "Error Summary",
    ]
    assert [c.value for c in wb["Status Changes"][1]] == ["Account", "Status", "From", "To", "Duration"]


def test_running_runs_are_excluded_from_the_report(db_session):
    site = _seed(db_session)
    buffer = build_site_report(db_session, site)
    wb = load_workbook(buffer)

    statuses = [row[4].value for row in wb["Check Runs"].iter_rows(min_row=2)]
    assert "RUNNING" not in statuses
    assert len(statuses) == 5  # 6 seeded runs minus the one still running


def test_status_changes_collapses_consecutive_same_status_runs(db_session):
    site = _seed(db_session)
    buffer = build_site_report(db_session, site)
    wb = load_workbook(buffer)

    rows = list(wb["Status Changes"].iter_rows(min_row=2, values_only=True))
    # success(x2), fail(x2), success(x1) -> 3 collapsed periods, not 5 individual rows
    assert len(rows) == 3
    assert [r[1] for r in rows] == ["SUCCESS", "FAIL", "SUCCESS"]


def test_report_filename_sanitizes_the_site_name(db_session):
    site = _seed(db_session)
    name = report_filename(site)
    assert name.startswith("Kiran_Chenna_Prod_report_")
    assert name.endswith(".xlsx")
    assert " " not in name
