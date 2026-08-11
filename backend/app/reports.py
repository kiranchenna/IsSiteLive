import io
from collections import defaultdict
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app import models

HEADER_FONT = Font(bold=True)
UP_FILL = PatternFill(start_color="D9F2E6", end_color="D9F2E6", fill_type="solid")
DOWN_FILL = PatternFill(start_color="FBDCDC", end_color="FBDCDC", fill_type="solid")


def _fmt(dt) -> str:
    return dt.isoformat(sep=" ", timespec="seconds") if dt else ""


def _status_fill(status: models.RunStatus) -> PatternFill:
    return UP_FILL if status == models.RunStatus.success else DOWN_FILL


def build_site_report(
    db: Session,
    site: models.Site,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    status_filter: Optional[models.RunStatus] = None,
) -> io.BytesIO:
    """An .xlsx workbook with the raw check log and a collapsed up/down timeline per account.

    start_date/end_date filter on started_at (inclusive); status_filter narrows the raw
    "Check Runs" log to just success or just fail rows. It deliberately does NOT narrow
    "Status Changes": collapsing consecutive same-status runs into periods only makes sense
    against the full timeline -- filtering out (say) every success run first would merge
    surrounding fail runs into one falsely-continuous outage instead of the real up/down/up
    pattern that actually happened.
    """
    query = db.query(models.CheckRun).filter(
        models.CheckRun.site_id == site.id, models.CheckRun.status != models.RunStatus.running
    )
    if start_date is not None:
        query = query.filter(models.CheckRun.started_at >= start_date)
    if end_date is not None:
        query = query.filter(models.CheckRun.started_at <= end_date)
    timeline_runs = query.order_by(models.CheckRun.started_at.asc()).all()
    runs = [r for r in timeline_runs if status_filter is None or r.status == status_filter]
    accounts = {a.id: a.label for a in db.query(models.Account).filter(models.Account.site_id == site.id).all()}

    wb = Workbook()

    runs_sheet = wb.active
    runs_sheet.title = "Check Runs"
    runs_sheet.append(["Started At", "Finished At", "Duration (s)", "Account", "Status", "Error Summary"])
    for cell in runs_sheet[1]:
        cell.font = HEADER_FONT
    for run in runs:
        runs_sheet.append(
            [
                _fmt(run.started_at),
                _fmt(run.finished_at),
                round(run.duration_ms / 1000, 1) if run.duration_ms is not None else "",
                accounts.get(run.account_id, f"#{run.account_id}"),
                run.status.value.upper(),
                run.error_summary or "",
            ]
        )
        status_cell = runs_sheet.cell(row=runs_sheet.max_row, column=5)
        status_cell.fill = _status_fill(run.status)
    for i, width in enumerate([20, 20, 12, 16, 10, 70], start=1):
        runs_sheet.column_dimensions[get_column_letter(i)].width = width
    runs_sheet.freeze_panes = "A2"

    changes_sheet = wb.create_sheet("Status Changes")
    changes_sheet.append(["Account", "Status", "From", "To", "Duration"])
    for cell in changes_sheet[1]:
        cell.font = HEADER_FONT

    by_account: dict[int, list[models.CheckRun]] = defaultdict(list)
    for run in timeline_runs:
        by_account[run.account_id].append(run)

    for account_id, account_runs in by_account.items():
        periods: list[dict] = []
        for run in account_runs:
            end = run.finished_at or run.started_at
            if periods and periods[-1]["status"] == run.status:
                periods[-1]["end"] = end
            else:
                periods.append({"status": run.status, "start": run.started_at, "end": end})
        for period in periods:
            changes_sheet.append(
                [
                    accounts.get(account_id, f"#{account_id}"),
                    period["status"].value.upper(),
                    _fmt(period["start"]),
                    _fmt(period["end"]),
                    str(period["end"] - period["start"]),
                ]
            )
            status_cell = changes_sheet.cell(row=changes_sheet.max_row, column=2)
            status_cell.fill = _status_fill(period["status"])
    for i, width in enumerate([16, 10, 20, 20, 16], start=1):
        changes_sheet.column_dimensions[get_column_letter(i)].width = width
    changes_sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def report_filename(site: models.Site, status_filter: Optional[models.RunStatus] = None) -> str:
    safe_name = "".join(c if c.isalnum() else "_" for c in site.name).strip("_") or f"site-{site.id}"
    suffix = f"_{status_filter.value}" if status_filter is not None else ""
    return f"{safe_name}_report{suffix}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
